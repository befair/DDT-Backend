import datetime

from api.models import AppUser, Client, Pallet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

ITA_MONTH = {
    1: "Gennaio",
    2: "Febbraio",
    3: "Marzo",
    4: "Aprile",
    5: "Maggio",
    6: "Giugno",
    7: "Luglio",
    8: "Agosto",
    9: "Settembre",
    10: "Ottobre",
    11: "Novembre",
    12: "Dicembre"
}

ITA_DAY = {
    0: "Lunedì",
    1: "Martedì",
    2: "Mercoledì",
    3: "Giovedì",
    4: "Venerdì",
    5: "Sabato",
    6: "Domenica"
}

DAY_START = datetime.time(6, 0, 0)
DAY_END = datetime.time(19, 0, 0)


def export_xlsx(ddts, filename):
    # Open XLSX file
    wb = Workbook()
    ws = wb.active

    data = {i: [] for i in range(1, 54)}

    # XLXS styles
    font = Font(name="Liberation Sans")
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left')
    border = Side(border_style="thin", color="000000")

    for ddt in ddts:
        index = 1
        op = AppUser.objects.get(pk=ddt.operator.pk)
        client = Client.objects.get(pk=ddt.client.pk)

        if ddt.time < DAY_START:
            day_offset = -1
        elif ddt.time > DAY_END:
            day_offset = 1
        else:
            day_offset = 0

        # Set DDT general info
        col = [
            f"{op.first_name} {op.last_name}",
            datetime.datetime.now(),
            ITA_DAY[ddt.date.weekday()+day_offset],
            ITA_MONTH[ddt.date.month],
            client.corporate_name,
            ddt.pk,
            ddt.serial,
            ddt.date
        ]

        for value in col:
            data[index].append(value)
            index += 1

        # Retrive all pallets for the current DDT
        for kind in Pallet.KIND:
            try:
                pallet = Pallet.objects.get(ddt=ddt.pk, kind=kind[0])
                for value in [pallet.received, pallet.returned, pallet.moved]:
                    data[index].append(value)
                    index += 1
            except Pallet.DoesNotExist:
                for _ in range(3):
                    data[index].append(0)
                    index += 1

            data[index].append('')
            index += 2

    # Write data to XLSX
    for i, row in enumerate(data, 1):
        for j, value in enumerate(data[row], 1):
            p = f"{index_to_xls(j)}{i}"
            ws[p] = value

            # Fix document style
            ws[p].alignment = left
            ws[p].font = font

    # Insert new column for labels
    ws.insert_cols(1)

    i = 9
    for kind in Pallet.KIND:
        # Write to file
        ws[f'A{i}'] = kind[1]

        # Styling
        for col in ws.columns:
            length = max(len(str(cell.value)) for cell in col)
            ws[f"{index_to_xls(col[0].column)}{i}"].border = Border(top=border)
            ws[f"{index_to_xls(col[0].column)}{i+4}"].border = Border(bottom=border)

        # Styling
        for p in range(i, i+4):
            ws[f'A{p}'].alignment = center
            ws[f'A{p}'].font = Font(name=font.name, bold=True)
            ws[f'A{p}'].border = Border(top=border,
                                        left=border,
                                        right=border,
                                        bottom=border)

        # Merge all cells of 'kind'
        ws.merge_cells(f'A{i}:A{i+4}')

        # Go to next position
        i += 5

    # Set correct colums width
    for col in ws.columns:
        length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[index_to_xls(col[0].column)].width = length

    # Save the file
    wb.save(filename)


def index_to_xls(col):
    """Convert integer to XLSX column (A,B,C ... AA,AB,AC ...)"""
    result = []

    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = chr(rem+65)
    return ''.join(result)
